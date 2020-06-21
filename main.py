import argparse
import datetime
from collections import namedtuple
import db
from openpyxl import load_workbook

# 1 квартала(кв)  = 1 Январь
# 2 квартала(кв)  = 1 Апреля
# 3 квартала(кв)  = 1 Июль
# 4 квартала(кв)  = 1 Октябрь
dict_quarters = {1: ('01', '01'), 2: ('01', '04'), 3: ('01', '07'), 4: ('01', '10')}
separators = {'dot': '.', 'dash': '-', 'zero_data': '#N/A'}


def get_headers():
    for row in sheet.iter_rows(min_row=3,
                               max_row=3,
                               min_col=1,
                               max_col=16,
                               values_only=True):
        # lst_rows =['№ п/п', 'Улица', '№ дома', 'Корпус',...]
        return list(row)


def trash_remover(list_headers):
    # list_headers =['№ п/п',... 'Корпус'...,'Место установки существующего ПУ\n'...]
    for idx, header_name in enumerate(list_headers):
        for element in header_name:
            if element.endswith('\n'):  # 'Место установки существующего ПУ\n'
                new_header_name = header_name.replace('\n', '')
                list_headers[idx] = new_header_name
    return list_headers


def analyze_quarter(row_lst, lst_to_return):
    # lst_with_quarter = [01, 'кв', '2029']
    lst_with_quarter = row_lst[10].split()

    for quarter in dict_quarters:
        if int(lst_with_quarter[0]) == quarter:
            lst_with_quarter[0] = dict_quarters[quarter][0]
            lst_with_quarter[1] = dict_quarters[quarter][1]
            row_lst[10] = separators['dash'].join(lst_with_quarter[::-1])
            lst_to_return = row_lst
    return lst_to_return


def analyze_date_check(row_lst):
    date_check = row_lst[10]
    lst_to_return = []
    if issubclass(type(date_check), str):
        # пристутствует слово квартал(кв)
        if 'кв' in date_check:
            result = analyze_quarter(row_lst, lst_to_return)
            return result
        else:
            for delimeter in date_check:
                if delimeter == separators['dot']:
                    # row_lst[10] = 19.01.2029
                    data = date_check.split(separators['dot'])
                    reversed_data = data[::-1]
                    row_lst[10] = separators['dash'].join(reversed_data)
                    lst_to_return = row_lst
                    break

                elif delimeter == separators['dash']:
                    # row_lst[10] = '-'
                    if date_check == separators['dash']:
                        row_lst[10] = '1801-01-01'
                        lst_to_return = row_lst
                        break
                    else:
                        # row_lst[10]=  01-2019
                        data = date_check.split(separators['dash'])
                        data.insert(0, dict_quarters[1][0])
                        row_lst[10] = separators['dash'].join(data[::-1])
                        lst_to_return = row_lst
                        break

                elif date_check == separators['zero_data']:
                    row_lst[10] = '1801-01-01'
                    lst_to_return = row_lst
                    break

            return lst_to_return


def analyze_date_of_mount(row_lst):
    lst_to_return = None
    date_of_mount = row_lst[9]
    if issubclass(type(date_of_mount), datetime.datetime):
        row_lst[9] = date_of_mount.strftime("%Y-%m-%d")
        checked_row = analyze_date_check(row_lst)
        lst_to_return = checked_row

    elif issubclass(type(date_of_mount), str):
        if date_of_mount == '-' or date_of_mount == 'не задано':
            row_lst[9] = '1801-01-01'
            checked_row = analyze_date_check(row_lst)
            lst_to_return = checked_row

    return lst_to_return


def get_row():
    corrected_rows_lst = []
    for row in sheet.iter_rows(min_row=4,
                               max_row=2732,
                               min_col=1,
                               max_col=16,
                               values_only=True):
        corrected_rows_lst.append(analyze_date_of_mount(list(row))[1::])
    return corrected_rows_lst


def insert_into_tables(lst_rows):
    # в исходных данных нет 1444 строчки
    for i in lst_rows:
        building_data = [i[0], i[1], i[2], i[3], i[7]]
        stock_info = [i[4], i[5], i[6], i[7], i[8], i[9], i[10], i[11], i[12], i[13], i[14]]
        Buildings = namedtuple('Buildings', 'street, building_no , building_corps, construction_no, device_no')

        Stock = namedtuple('Stock', '''tariff,
                                        coefficient,
                                        mount_place,
                                        device_no,
                                        date_of_mount,date_check,
                                        transform_coeff,
                                        type_current,
                                        no_phase_a,
                                        no_phase_b,
                                        no_phase_c ''')

        table_building = Buildings._make(building_data)
        table_stock_info = Stock._make(stock_info)

        db.insert_building_info(con,
                                table_building.street,
                                table_building.building_no,
                                table_building.building_corps,
                                table_building.construction_no,
                                table_building.device_no)

        db.insert_stock_info(con,
                             table_stock_info.tariff,
                             table_stock_info.coefficient,
                             table_stock_info.mount_place,
                             table_stock_info.device_no,
                             table_stock_info.date_of_mount,
                             table_stock_info.date_check,
                             table_stock_info.transform_coeff,
                             table_stock_info.type_current,
                             table_stock_info.no_phase_a,
                             table_stock_info.no_phase_b,
                             table_stock_info.no_phase_c)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='xlxs analyzer')
    parser.add_argument('-f', '--file', help="Try this: python3 main.py -f=test_data.xlsx", type=str,
                        default='test_data.xlsx')
    args = parser.parse_args()
    file_xlxs = args.file

    workbook = load_workbook(file_xlxs)
    sheet = workbook.active
    list_headers = trash_remover(get_headers())
    lst_rows = get_row()

    con = db.connection_db()
    db.create_tables(con)

    insert_into_tables(lst_rows)
